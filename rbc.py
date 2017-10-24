from dateutil.parser import parse
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import argparse
import os


def is_date(string):
    try:
        parse(string)
        return True
    except ValueError:
        return False

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False

def extract(strs):
    balance = None
    withdrawals = None
    desc = None

    for s in strs:
        s = s.split()
        left,right = 0,0

        if is_number(s[-1]) and is_number(s[-2]):
            balance = s[-1]
            withdrawals = s[-2]
            right = len(s) - 2
        elif s[-1].isdigit():
            balance = balance #prev
            withdrawals = s[-1]
            right = len(s) - 1

        if s[0].isdigit():
            date = s[0] + s[1]
            left = 1

        #desc = s[left:right]
        #if desc[-1].isdigit(): # 2 liners
        #    pass

        print(balance,withdrawals,desc)


def _____parse_excel(l):
    i = 0
    while i<len(l):
        words = l[i].split()
        if is_date(words[0]):
            date = ' '.join(words[:2])
            if words[-1][0] == '$' or words[-1][0] == '-': # if 1 liner
                price = words[-1]
                location =  ' '.join(words[4:-1])
                i+=1
            else:
                location =  ' '.join(words[4:])
                i+=2
                price = l[i]
                i+=1
            if price[0] == "(" or price[0] == '-':
                credit = True
            else:
                credit = False
            print("data: {}, location: {}, price: {}".format(
                  date,location, price))
        else:
            print("error", l[i])
            i+=1

def  _parse_worksheet(in_ws, out_ws, verbose):
    i = 1
    row = 2
    while i < in_ws.max_row:
        words = in_ws.cell(column=1, row=i).value.split()
        # Transaction usually starts w date
        if is_date(words[0]):
            date = ' '.join(words[:2])

            # If 3 liner
            if words[-1][0] == '$' or words[-1][0] == '-':  # if 1 liner
                price = words[-1]
                description = ' '.join(words[4:-1])
            # If 1 liner
            else:
                description = ' '.join(words[4:])
                i += 2
                price = in_ws.cell(column=1, row=i).value
       # Case when new month starts
        elif words[0] == 'NEW':
            description = "NEW BALANCE"
            date = ""
            price = words[-1]
        # Unable to parse
        else:
            print("error", in_ws.cell(column=1, row=i).value)
            i += 1
            continue
        if verbose:
            print("data: {}, description: {}, price: {}".format(
                date, description, price))
        out_ws.cell(column=1, row=row).value = date
        out_ws.cell(column=2, row=row).value = description
        out_ws.cell(column=3, row=row).value = price
        row += 1
        i += 1


def excel(wb_path, prefix='rbc-parser-', new=False, verbose=False, confirm=False):
    headers = ['Date', 'Description', 'Price']

    wb = load_workbook(wb_path)
    for ws in wb.worksheets:
        if ws.title.startswith("rbc-parser"):
            continue
        if confirm:
            resp = input("parse <{}> ?".format(ws.title))
            if resp.lower() == 'y' or resp.lower() == 'yes':
                out_name = prefix + ws.title
                out_ws = wb.create_sheet(title=out_name)
                out_ws['A1'].value = headers[0]
                out_ws['B1'].value = headers[1]
                out_ws['C1'].value = headers[2]
                _parse_worksheet(ws,out_ws,verbose)
                print()
            else:
                continue
    if new:
        wb_path = os.path.join(os.path.dirname(wb_path), prefix+os.path.basename(wb_path))
        wb.save(wb_path)
        print
    else:
        wb.save(wb_path)
    print("saved workbook at {}".format(wb_path))



if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Parse rbc statements', epilog='not affiliated w/ rbc')
    parser.add_argument('wb_path', action="store", help="path to the excel workbook")
    parser.add_argument('-v', '--verbose', help='verbose',default=False, dest='verbose', action="store_true")
    parser.add_argument('-c', '--confirmation' , help='asks for confirmation before parsing worksheet if true',
                        default=False, dest='confirm', action="store_true")
    parser.add_argument('-p', '--prefix', help='assign prefix for wb/ws names', default='rbc-parser-',
                        dest='prefix', action='store')
    parser.add_argument('-n', '--new', help='create new workbook instead of saving existing one', default=False, dest='new',
                        action="store_true")


    options = parser.parse_args()
    excel(**vars(options))



